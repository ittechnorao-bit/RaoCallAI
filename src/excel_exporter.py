from pathlib import Path
from openpyxl import Workbook, load_workbook

EXCEL_DIR = Path(r"D:\RaoCallAI\output\excel")
EXCEL_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_FILE = EXCEL_DIR / "Transcript.xlsx"


def save_transcript(file_name, original_text, english_text):

    if EXCEL_FILE.exists():

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    else:

        wb = Workbook()
        ws = wb.active

        ws.append([
            "File Name",
            "Original Transcript",
            "English Translation"
        ])

    ws.append([
        file_name,
        original_text,
        english_text
    ])

    wb.save(EXCEL_FILE)